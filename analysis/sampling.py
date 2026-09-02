"""Build the blind double-coding workbook for inter-rater reliability (I2)."""
import pandas as pd
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import config

_INSTRUCTIONS = [
    "INTER-RATER RELIABILITY CODING — ESG SYSTEMATIC REVIEW",
    "",
    "Purpose: Reviewers 1 and 2 both require an inter-rater reliability statistic",
    "for the 0-5 coding scheme. This workbook produces it.",
    "",
    "HOW TO USE",
    "1. Each expert codes INDEPENDENTLY. Expert A must not see Expert B's sheet,",
    "   and neither expert may consult the other while coding. Independence is the",
    "   whole point; conferring invalidates the statistic.",
    "2. Code every one of the 44 sub-dimensions for every article listed.",
    "3. Use only whole numbers 0-5. See the Codebook sheet for the anchors.",
    "4. Leave nothing blank. If an article does not address a sub-dimension, enter 0.",
    "5. If you recognise an article and recall the score it originally received,",
    "   write RECALL in the Notes column for that row. This is not a problem - it is",
    "   disclosed as a limitation of re-using the original coders - but it must be",
    "   recorded, not silently absorbed into the score.",
    "6. Return the file with both sheets completed.",
    "",
    "The articles below are a stratified random sample of the 1,026-article corpus,",
    "drawn proportionally by publication year with a fixed seed.",
]


def stratified_subsample(df, frac: float = 0.10):
    """Proportional-by-year random sample, reproducible under config.SEED."""
    parts = [g.sample(n=max(1, round(len(g) * frac)), random_state=config.SEED)
             for _, g in df.groupby("year")]
    sample = pd.concat(parts)
    return sample.sort_values("sr_no").reset_index(drop=True)


def write_coding_workbook(sample, path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Instructions"
    ws.column_dimensions["A"].width = 90
    for i, line in enumerate(_INSTRUCTIONS, start=1):
        ws.cell(row=i, column=1, value=line)
    ws["A1"].font = Font(bold=True, size=13)

    cb = wb.create_sheet("Codebook")
    cb.append(["CODE", "DIMENSION", "SUB-DIMENSION"])
    for code, label in config.SUBDIMENSIONS.items():
        cb.append([code, config.DIMENSIONS[code[0]], label])
    cb.append([])
    cb.append(["SCORE", "ANCHOR"])
    for score, anchor in config.SCALE_ANCHORS.items():
        cb.append([score, anchor])
    for cell in cb[1]:
        cell.font = Font(bold=True)
    cb.column_dimensions["A"].width = 10
    cb.column_dimensions["B"].width = 34
    cb.column_dimensions["C"].width = 56

    for expert in ("Expert A", "Expert B"):
        sh = wb.create_sheet(expert)
        sh.append(["SR.NO", "YEAR", "TITLE", "ABSTRACT"] + config.SCORE_COLS + ["Notes"])
        for cell in sh[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(textRotation=90, vertical="bottom")
        for _, r in sample.iterrows():
            # Handle pd.NA for missing title by converting to empty string
            title = r["title"] if pd.notna(r["title"]) else ""
            abstract = r["abstract"] if pd.notna(r["abstract"]) else ""
            sh.append([int(r["sr_no"]), int(r["year"]), title, abstract])
        sh.freeze_panes = "E2"
        sh.column_dimensions["A"].width = 8
        sh.column_dimensions["B"].width = 7
        sh.column_dimensions["C"].width = 60
        sh.column_dimensions["D"].width = 80
        for j in range(5, 5 + len(config.SCORE_COLS)):
            sh.column_dimensions[get_column_letter(j)].width = 6
        notes_col = get_column_letter(5 + len(config.SCORE_COLS))
        sh.column_dimensions[notes_col].width = 40

        # Add data validation for score cells
        dv = DataValidation(
            type="whole", operator="between", formula1=0, formula2=5, allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid score",
            error="Enter a whole number from 0 to 5. See the Codebook sheet for the anchors.",
        )
        sh.add_data_validation(dv)
        last_col = get_column_letter(4 + len(config.SCORE_COLS))
        dv.add(f"E2:{last_col}{1 + len(sample)}")

    wb.save(path)
    return path
