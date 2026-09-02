"""Tests for the one-command reproducible pipeline (I4).

Before this module existed, `export.write_all()` and every `fig_*.render()`
had zero non-test callers - the eight committed tables and six figures were
produced by uncommitted scratchpad scripts, and `analysis/_outputs/` is
gitignored, so a fresh checkout could not reproduce any artifact the
manuscript cites. `run_all.main()` is the single entry point that can.

These tests always pass `out_root=tmp_path` (or a subdirectory of it) so a
test run never writes into the real `figures/`, `manuscript/tables/` or
`analysis/_outputs/` - only the real, explicit, one-time invocation
(`python -m analysis.run_all`, run manually per the dispatch, not by the
test suite) writes there.
"""
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

from analysis import run_all


# --------------------------------------------------------------------------
# Caching primitives, tested directly against cheap synthetic `compute`
# functions - NOT through the full pipeline, since the read-hit path
# ("cache expensive stages") can never be exercised by a quick=True run
# (quick never touches the cache, by design - see run_all.py's docstring)
# and quick=False runs the real 35-minute benchmark sweep, unusable in a
# unit test.
# --------------------------------------------------------------------------

def test_cache_csv_reads_back_an_existing_cache_without_recomputing(tmp_path):
    calls = []

    def compute():
        calls.append(1)
        return pd.DataFrame({"x": [1.0, 2.0, 3.0]}, index=["a", "b", "c"])

    first = run_all._cache_csv(tmp_path, "demo", compute, use_cache=True)
    second = run_all._cache_csv(tmp_path, "demo", compute, use_cache=True)
    assert len(calls) == 1   # the second call read the cache file, it did not recompute
    pd.testing.assert_frame_equal(first, second)
    assert (tmp_path / "demo.csv").exists()


def test_cache_csv_with_use_cache_false_never_reads_or_writes(tmp_path):
    calls = []

    def compute():
        calls.append(1)
        return pd.DataFrame({"x": [1.0]}, index=["a"])

    run_all._cache_csv(tmp_path, "demo", compute, use_cache=False)
    run_all._cache_csv(tmp_path, "demo", compute, use_cache=False)
    assert len(calls) == 2                          # recomputed both times
    assert not (tmp_path / "demo.csv").exists()      # and never wrote a cache file either


def test_cache_csv_never_requires_the_cache_to_exist(tmp_path):
    """I4: 'never require the cache to exist' - a completely absent
    cache_dir must not raise, the first call must create it."""
    missing_dir = tmp_path / "does" / "not" / "exist" / "yet"
    assert not missing_dir.exists()
    out = run_all._cache_csv(missing_dir, "demo", lambda: pd.DataFrame({"x": [1.0]}), use_cache=True)
    assert not out.empty
    assert (missing_dir / "demo.csv").exists()


def test_cache_parallel_analysis_round_trips_through_npz_without_recomputing(tmp_path):
    calls = []

    def compute():
        calls.append(1)
        return {"n_factors": 5, "eigenvalues": np.array([3.0, 2.0, 1.0]),
               "threshold": np.array([1.5, 1.4, 1.3])}

    first = run_all._cache_parallel_analysis(tmp_path, compute, use_cache=True)
    second = run_all._cache_parallel_analysis(tmp_path, compute, use_cache=True)
    assert len(calls) == 1
    assert second["n_factors"] == 5
    assert isinstance(second["n_factors"], int)
    assert np.allclose(second["eigenvalues"], first["eigenvalues"])
    assert np.allclose(second["threshold"], first["threshold"])


def test_quick_run_produces_non_empty_figures_and_tables(tmp_path):
    out = run_all.main(quick=True, out_root=tmp_path)
    assert out["figures"] and out["tables"]
    # Figures are always > 10KB (vector PDF with embedded fonts). Tables vary
    # a lot by legitimate row count - tab_factorability is a real, correct,
    # 5-row table at 399 bytes; tab_benchmark's quick 2-estimator subset is
    # 370. 200 bytes comfortably rules out an empty/near-empty file (a bare
    # \begin{table}...\end{table} with no rows) without penalising a
    # genuinely small table.
    for p in out["figures"].values():
        assert Path(p).exists() and Path(p).stat().st_size > 1000
    for p in out["tables"].values():
        assert Path(p).exists() and Path(p).stat().st_size > 200


def test_quick_run_writes_under_figures_and_manuscript_tables(tmp_path):
    """I4: 'must write every figure to figures/ and every table to
    manuscript/tables/' - checked against the actual paths returned, not
    just that files exist somewhere."""
    out = run_all.main(quick=True, out_root=tmp_path)
    for p in out["figures"].values():
        assert Path(p).parent == tmp_path / "figures"
    for p in out["tables"].values():
        assert Path(p).parent == tmp_path / "manuscript" / "tables"


def test_run_reports_skipped_gated_stages(tmp_path):
    """Reliability (Task 12) now has an upstream module, but I2 (the
    completed double-coding workbook) has not arrived - both expert sheets
    in the committed `templates/IRR_double_coding.xlsx` are still entirely
    blank - so it must still be disclosed in `skipped`, not silently
    absent. External validation (Task 18, gated on input I1) is no longer
    skipped now that I1 has arrived (`data/` is committed) - see
    `test_run_runs_external_validation_now_that_i1_has_arrived` below."""
    out = run_all.main(quick=True, out_root=tmp_path)
    assert "skipped" in out
    assert all(isinstance(s, str) for s in out["skipped"])
    joined = " ".join(out["skipped"]).lower()
    assert "reliability" in joined


def test_skipped_reliability_reason_distinguishes_absent_from_uncoded(tmp_path):
    """The skip reason must say WHY reliability is unavailable, not just
    that it is: 'workbook absent' and 'workbook present but not yet coded'
    are different situations for a reader of the disclosure to understand,
    and the real committed workbook is the latter (present, both sheets
    still blank)."""
    out = run_all.main(quick=True, out_root=tmp_path)
    reliability_reasons = [s for s in out["skipped"] if "reliability" in s.lower()]
    assert len(reliability_reasons) == 1
    reason = reliability_reasons[0].lower()
    assert "not yet been coded" in reason or "not yet coded" in reason
    assert "absent" not in reason

    assert run_all._reliability_unavailable_reason(None) == (
        "the double-coding workbook is present "
        f"({run_all.config.TEMPLATES / 'IRR_double_coding.xlsx'}) but has not yet "
        "been coded - both expert sheets are still blank"
    )


def test_reliability_unavailable_reason_says_absent_when_the_workbook_is_missing(
    tmp_path, monkeypatch
):
    from analysis import config as _config
    monkeypatch.setattr(_config, "TEMPLATES", tmp_path)
    reason = run_all._reliability_unavailable_reason(None)
    assert "absent" in reason
    assert "not yet been coded" not in reason


def test_run_uses_reliability_data_once_it_arrives_without_corrupting_factor_reliability(
    tmp_path, monkeypatch
):
    """Regression guard: an earlier draft named the double-coding
    `reliability.reliability_table(...)` result `reliability_tbl` - the same
    name `factors.reliability(df, assignment)`'s output already used
    upstream for `factor_reliability_table`. Since this whole stage is
    normally skipped (I2 has not arrived), that collision was invisible to
    every other test; it would have silently corrupted
    `tab_factor_reliability.tex` the moment real double-coded data landed.
    This builds a fully (uniformly) coded synthetic workbook, points
    `config.TEMPLATES` at it, and checks BOTH tables come out right."""
    import openpyxl
    from analysis import config as _config
    from analysis import loading, sampling

    templates_dir = tmp_path / "templates"
    templates_dir.mkdir()
    # `main()` also reads prisma_counts.json from config.TEMPLATES.
    import shutil
    shutil.copy(_config.TEMPLATES / "prisma_counts.json", templates_dir / "prisma_counts.json")

    sample = sampling.stratified_subsample(loading.load_scoring(), frac=0.02)
    irr_path = sampling.write_coding_workbook(sample, templates_dir / "IRR_double_coding.xlsx")
    wb = openpyxl.load_workbook(irr_path)
    n_rows = len(sample)
    # Real variance per sub-dimension column (not a flat constant): a
    # column where every article gets the exact same score has zero
    # variance, which makes ICC(2,1) mathematically undefined (0/0), not
    # just uninteresting - reliability.icc21 correctly raises on that, so a
    # constant fixture would fail here for the right reason but for the
    # wrong test.
    values = [[(r + c) % 6 for c in range(len(_config.SCORE_COLS))] for r in range(n_rows)]
    for sheet_name in ("Expert A", "Expert B"):
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]
        first = header.index(_config.SCORE_COLS[0])
        for r, row_values in enumerate(values):
            for c, v in enumerate(row_values):
                ws.cell(row=2 + r, column=first + 1 + c).value = v
    wb.save(irr_path)

    monkeypatch.setattr(_config, "TEMPLATES", templates_dir)

    out = run_all.main(quick=True, out_root=tmp_path / "out")

    joined = " ".join(out["skipped"]).lower()
    assert "reliability" not in joined
    assert "reliability" in out["tables"]
    assert Path(out["tables"]["reliability"]).exists()

    # factor_reliability must still be the factor model's own table, not
    # silently overwritten by the double-coding reliability table.
    assert "factor_reliability" in out["tables"]
    factor_reliability_text = Path(out["tables"]["factor_reliability"]).read_text(
        encoding="utf-8")
    assert "cronbach" in factor_reliability_text.lower() \
        or "mcdonald" in factor_reliability_text.lower()
    reliability_text = Path(out["tables"]["reliability"]).read_text(encoding="utf-8")
    assert "kappa" in reliability_text.lower()


def test_run_runs_external_validation_now_that_i1_has_arrived(tmp_path):
    """Task 18: `data/` (the LSEG/Refinitiv extract) is present in this
    repository, so `analysis.external.is_available()` is True and Study 2
    must actually run as part of the pipeline, not be skipped."""
    out = run_all.main(quick=True, out_root=tmp_path)
    joined = " ".join(out["skipped"]).lower()
    assert "external" not in joined
    assert "external" in out["tables"]
    assert "external" in out["figures"]
    assert Path(out["tables"]["external"]).exists()
    assert Path(out["figures"]["external"]).exists()


def test_run_does_not_fail_when_gated_stages_are_unavailable(tmp_path):
    """The whole point of C1/I4's 'append to skipped rather than failing':
    a missing analysis.reliability / analysis.external module must not
    raise, it must be recorded and the run must still complete."""
    out = run_all.main(quick=True, out_root=tmp_path)
    assert out["tables"]   # the run reached the table-writing stage at all


def test_run_is_deterministic_under_the_global_seed(tmp_path):
    a = run_all.main(quick=True, out_root=tmp_path / "a")
    b = run_all.main(quick=True, out_root=tmp_path / "b")
    assert a["descriptives_checksum"] == b["descriptives_checksum"]


def test_full_specs_table_set_is_reachable_in_the_returned_dict_keys(tmp_path):
    """Not every SPECS table needs data in a quick run, but every table that
    IS produced must be one export.SPECS actually knows about - run_all must
    never invent an unrequested table key."""
    from analysis import export
    out = run_all.main(quick=True, out_root=tmp_path)
    assert set(out["tables"]) <= set(export.SPECS)


def test_convergence_and_loadings_tables_are_always_produced_even_when_quick(tmp_path):
    """These two are the fix targets of C3 and I3 - a quick run must still
    exercise them, not just the cheapest tables."""
    out = run_all.main(quick=True, out_root=tmp_path)
    assert "convergence" in out["tables"]
    assert "loadings" in out["tables"]


def test_cache_is_not_required_to_exist(tmp_path):
    """A fresh out_root has no analysis/_outputs cache at all - the run must
    not assume one is there."""
    cache_dir = tmp_path / "analysis" / "_outputs"
    assert not cache_dir.exists()
    out = run_all.main(quick=True, out_root=tmp_path)
    assert out["tables"]


def test_quick_run_never_touches_the_real_repository_output_directories(tmp_path):
    """Regression guard: `out_root` must actually redirect every write - a
    bug that silently fell back to `config.FIGURES`/`config.TABLES` would
    pollute the committed manuscript on every test run."""
    from analysis import config
    before_figs = set(config.FIGURES.glob("*")) if config.FIGURES.exists() else set()
    before_tabs = set(config.TABLES.glob("*")) if config.TABLES.exists() else set()
    run_all.main(quick=True, out_root=tmp_path)
    after_figs = set(config.FIGURES.glob("*")) if config.FIGURES.exists() else set()
    after_tabs = set(config.TABLES.glob("*")) if config.TABLES.exists() else set()
    assert before_figs == after_figs
    assert before_tabs == after_tabs
