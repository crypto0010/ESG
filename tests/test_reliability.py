"""Tests for inter-rater reliability (Task 12, R1.3 / R2.2), gated on
author input I2 -- the completed `templates/IRR_double_coding.xlsx`.

Neither expert sheet has been coded yet (both are entirely blank), so this
module is written and tested entirely against synthetic fixtures. The most
important test in this file is `test_load_refuses_the_real_blank_workbook`:
it checks against the REAL, committed template, and it is the guard against
`load_double_coding` ever silently manufacturing perfect agreement out of
blank cells.
"""
import openpyxl
import numpy as np
import pandas as pd
import pingouin as pg
import pytest

from analysis import config, loading, reliability, sampling

RNG = np.random.default_rng(config.SEED)
N_DIM_ROWS = len(config.DIMENSIONS)
N_SUB_ROWS = len(config.SCORE_COLS)


def _pair(n=120, noise=0.0):
    a = RNG.integers(0, 6, n).astype(float)
    b = a.copy()
    if noise:
        flip = RNG.random(n) < noise
        b[flip] = RNG.integers(0, 6, int(flip.sum()))
    return a, b


def _score_frame(n, rng=RNG):
    return pd.DataFrame({c: rng.integers(0, 6, n).astype(float) for c in config.SCORE_COLS})


def _code_sheet_uniform(ws, value):
    header = [c.value for c in ws[1]]
    first = header.index(config.SCORE_COLS[0])
    last = header.index(config.SCORE_COLS[-1])
    for row in ws.iter_rows(min_row=2):
        for cell in row[first:last + 1]:
            cell.value = value


def _score_cell_column(ws, offset=0):
    header = [c.value for c in ws[1]]
    return header.index(config.SCORE_COLS[0]) + 1 + offset


def _small_blank_workbook(tmp_path, name="irr.xlsx", frac=0.02):
    s = sampling.stratified_subsample(loading.load_scoring(), frac=frac)
    p = sampling.write_coding_workbook(s, tmp_path / name)
    return p, s


# --------------------------------------------------------------------------
# weighted_kappa
# --------------------------------------------------------------------------

def test_perfect_agreement_gives_kappa_one():
    a, _ = _pair()
    assert reliability.weighted_kappa(a, a) > 0.999


def test_noisy_agreement_lowers_kappa():
    a, b = _pair(noise=0.5)
    assert reliability.weighted_kappa(a, b) < reliability.weighted_kappa(a, a)


def test_kappa_near_zero_for_independent_coders():
    a = RNG.integers(0, 6, 500).astype(float)
    b = RNG.integers(0, 6, 500).astype(float)
    assert abs(reliability.weighted_kappa(a, b)) < 0.2


# --------------------------------------------------------------------------
# icc21
# --------------------------------------------------------------------------

def test_icc_is_one_for_identical_ratings():
    a, _ = _pair()
    assert reliability.icc21(a, a) > 0.99


def test_icc21_matches_pingouin_directly():
    """Verified against pingouin's own `intraclass_corr` (0.6.1), reading
    the same "ICC(A,1)" (absolute agreement, single rater) row `icc21`
    reads internally, on noisy (non-degenerate) data."""
    a, b = _pair(n=80, noise=0.4)
    n = len(a)
    long = pd.DataFrame({
        "target": list(range(n)) * 2,
        "rater": ["A"] * n + ["B"] * n,
        "score": np.concatenate([a, b]),
    })
    ref = pg.intraclass_corr(data=long, targets="target", raters="rater", ratings="score")
    expected = float(ref.loc[ref["Type"] == "ICC(A,1)", "ICC"].iloc[0])
    assert reliability.icc21(a, b) == pytest.approx(expected)


def test_systematic_offset_shows_icc_and_agreement_are_not_redundant():
    """A constant one-point offset (b = a + 1, kept in-range by drawing a
    from 0-4) means literally no cell can ever exactly match: percent
    agreement is exactly zero. ICC(2,1) still reflects the strong relative
    consistency, so it stays well above zero -- the two metrics diverge
    sharply on this pattern, which is exactly why both are reported rather
    than either alone."""
    n = 80
    a = RNG.integers(0, 5, n).astype(float)
    b = a + 1.0
    assert reliability.percent_agreement(a, b) == 0.0
    icc = reliability.icc21(a, b)
    assert icc > 0.5
    assert reliability.percent_agreement(a, b) < icc


# --------------------------------------------------------------------------
# percent_agreement
# --------------------------------------------------------------------------

def test_percent_agreement_is_a_proportion():
    a, b = _pair(noise=0.3)
    p = reliability.percent_agreement(a, b)
    assert 0.0 <= p <= 1.0
    assert reliability.percent_agreement(a, a) == 1.0


# --------------------------------------------------------------------------
# reliability_table
# --------------------------------------------------------------------------

def test_reliability_table_has_subdimension_dimension_and_overall_rows():
    n = 60
    ea = _score_frame(n)
    eb = ea.copy()
    t = reliability.reliability_table(ea, eb)
    assert len(t) == N_SUB_ROWS + N_DIM_ROWS + 1
    assert "OVERALL" in t.index
    for letter in config.DIMENSIONS:
        assert letter in t.index
    for code in config.SCORE_COLS:
        assert code in t.index
    assert list(t.columns) == ["n", "percent_agreement", "kappa", "icc"]
    assert (t["kappa"] > 0.99).all()
    assert (t["icc"] > 0.99).all()
    assert t.loc["OVERALL", "n"] == n * N_SUB_ROWS


# --------------------------------------------------------------------------
# bootstrap_overall
# --------------------------------------------------------------------------

def test_bootstrap_overall_resamples_articles_not_cells():
    """Halving the number of ARTICLES (not cells) must widen the interval.
    If the implementation resampled cells instead, halving the article
    count would barely move n_cells (44x fewer articles is still 44x more
    cells than "1 article"), and the interval would not visibly widen."""
    n = 40
    ea = _score_frame(n)
    flip = RNG.random((n, N_SUB_ROWS)) < 0.35
    noise = RNG.integers(0, 6, (n, N_SUB_ROWS)).astype(float)
    eb = ea.copy()
    eb.values[flip] = noise[flip]

    full = reliability.bootstrap_overall(ea, eb, n_boot=400)
    half = reliability.bootstrap_overall(
        ea.iloc[: n // 2].reset_index(drop=True),
        eb.iloc[: n // 2].reset_index(drop=True),
        n_boot=400,
    )
    assert (full["kappa_hi"] - full["kappa_lo"]) < (half["kappa_hi"] - half["kappa_lo"])
    assert (full["icc_hi"] - full["icc_lo"]) < (half["icc_hi"] - half["icc_lo"])


def test_bootstrap_overall_requires_at_least_two_articles():
    ea = _score_frame(1)
    with pytest.raises(ValueError, match="at least 2 articles"):
        reliability.bootstrap_overall(ea, ea, n_boot=50)


# --------------------------------------------------------------------------
# disagreement_profile
# --------------------------------------------------------------------------

def test_disagreement_profile_sorts_worst_item_first():
    n = 50
    ea = _score_frame(n)
    eb = ea.copy()
    worst_code = config.SCORE_COLS[3]
    eb[worst_code] = (ea[worst_code] + 3) % 6  # deliberately terrible on one item
    prof = reliability.disagreement_profile(ea, eb)
    assert list(prof.index) == sorted(prof.index, key=lambda c: -prof.loc[c, "mean_abs_diff"])
    assert prof.index[0] == worst_code
    assert prof.loc[worst_code, "exact_agreement"] < prof["exact_agreement"].median()


# --------------------------------------------------------------------------
# is_available / load_double_coding: file-backed behaviour
# --------------------------------------------------------------------------

def test_is_available_is_false_when_the_file_is_missing(tmp_path):
    assert reliability.is_available(tmp_path / "nope.xlsx") is False


def test_is_available_is_false_for_the_real_but_uncoded_workbook():
    """The committed templates/IRR_double_coding.xlsx exists but both
    expert sheets are entirely blank -- not yet coded, hence unavailable."""
    assert reliability.is_available() is False


def test_load_refuses_the_real_blank_workbook():
    """The generated template is blank; loading it must fail, not return
    zeros. This is the most important test in the module: returning zeros
    for blank cells would silently manufacture perfect agreement."""
    with pytest.raises(ValueError, match="not been coded") as exc:
        reliability.load_double_coding()
    message = str(exc.value)
    assert "Expert A" in message
    assert any(ch.isdigit() for ch in message)


def test_load_refuses_a_freshly_generated_blank_workbook(tmp_path):
    p, _ = _small_blank_workbook(tmp_path, "blank.xlsx")
    with pytest.raises(ValueError, match="not been coded"):
        reliability.load_double_coding(p)


def test_load_double_coding_returns_score_frames_when_fully_coded(tmp_path):
    p, s = _small_blank_workbook(tmp_path, "full.xlsx")
    wb = openpyxl.load_workbook(p)
    _code_sheet_uniform(wb["Expert A"], 3)
    _code_sheet_uniform(wb["Expert B"], 3)
    wb.save(p)

    a, b = reliability.load_double_coding(p)
    assert list(a.columns) == config.SCORE_COLS
    assert list(b.columns) == config.SCORE_COLS
    assert len(a) == len(s) == len(b)
    assert (a == 3).all().all()
    assert reliability.is_available(p) is True


def test_load_refuses_partial_coding(tmp_path):
    p, _ = _small_blank_workbook(tmp_path, "partial.xlsx")
    wb = openpyxl.load_workbook(p)
    _code_sheet_uniform(wb["Expert A"], 3)
    _code_sheet_uniform(wb["Expert B"], 3)
    # Blank out one cell that was just coded -> partial, not fully uncoded.
    # openpyxl's `cell(value=None)` is a "leave alone" sentinel, not "clear" -
    # setting `.value` directly is required to actually blank the cell.
    wb["Expert A"].cell(row=2, column=_score_cell_column(wb["Expert A"])).value = None
    wb.save(p)

    with pytest.raises(ValueError, match="partially coded") as exc:
        reliability.load_double_coding(p)
    assert "Expert A" in str(exc.value)
    assert "1 of" in str(exc.value)


def test_load_refuses_out_of_range_values(tmp_path):
    p, _ = _small_blank_workbook(tmp_path, "range.xlsx")
    wb = openpyxl.load_workbook(p)
    _code_sheet_uniform(wb["Expert A"], 3)
    _code_sheet_uniform(wb["Expert B"], 3)
    wb["Expert B"].cell(row=3, column=_score_cell_column(wb["Expert B"], offset=2), value=9)
    wb.save(p)

    with pytest.raises(ValueError, match="0-5") as exc:
        reliability.load_double_coding(p)
    assert "Expert B" in str(exc.value)


def test_load_refuses_mismatched_article_sets(tmp_path):
    p, _ = _small_blank_workbook(tmp_path, "mismatch.xlsx")
    wb = openpyxl.load_workbook(p)
    _code_sheet_uniform(wb["Expert A"], 3)
    _code_sheet_uniform(wb["Expert B"], 3)
    wb["Expert B"].cell(row=2, column=1, value=999999)  # SR.NO not present in Expert A
    wb.save(p)

    with pytest.raises(ValueError, match="different articles"):
        reliability.load_double_coding(p)


# --------------------------------------------------------------------------
# recall_flags
# --------------------------------------------------------------------------

def test_recall_flags_reads_notes_and_counts_flags(tmp_path):
    p, _ = _small_blank_workbook(tmp_path, "flags.xlsx")
    wb = openpyxl.load_workbook(p)
    ws_a, ws_b = wb["Expert A"], wb["Expert B"]
    notes_col = ws_a.max_column
    assert ws_a.cell(row=1, column=notes_col).value == "Notes"

    ws_a.cell(row=2, column=notes_col, value="I recall coding this one before.")
    ws_a.cell(row=3, column=notes_col, value="No issues.")
    ws_b.cell(row=2, column=notes_col, value="Nothing to flag.")
    wb.save(p)

    flags = reliability.recall_flags(p)
    assert flags["total_flags"] == 1
    assert flags["per_sheet"]["Expert A"]["count"] == 1
    assert flags["per_sheet"]["Expert B"]["count"] == 0
    sr_no_row2 = ws_a.cell(row=2, column=1).value
    assert flags["per_sheet"]["Expert A"]["articles"] == [sr_no_row2]
    assert flags["articles_flagged_by_either"] == [sr_no_row2]


def test_recall_flags_is_zero_on_the_untouched_real_workbook():
    """No coder has written anything yet -- zero flags, not an error, and
    not a fabricated count."""
    flags = reliability.recall_flags()
    assert flags["total_flags"] == 0
    assert flags["articles_flagged_by_either"] == []
