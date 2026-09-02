import openpyxl
from analysis import config, loading, sampling

DF = loading.load_scoring()


def test_subsample_is_about_ten_percent():
    s = sampling.stratified_subsample(DF, frac=0.10)
    assert 100 <= len(s) <= 110
    assert s["sr_no"].is_unique


def test_subsample_is_stratified_by_year():
    s = sampling.stratified_subsample(DF, frac=0.10)
    assert set(s["year"]) == set(DF["year"])
    # 2024 dominates the corpus, so it must dominate the sample too
    assert s["year"].value_counts().idxmax() == 2024


def test_subsample_is_reproducible():
    a = sampling.stratified_subsample(DF, frac=0.10)
    b = sampling.stratified_subsample(DF, frac=0.10)
    assert list(a["sr_no"]) == list(b["sr_no"])


def test_workbook_has_the_expected_sheets(tmp_path):
    s = sampling.stratified_subsample(DF, frac=0.10)
    p = sampling.write_coding_workbook(s, tmp_path / "irr.xlsx")
    wb = openpyxl.load_workbook(p)
    assert wb.sheetnames == ["Instructions", "Codebook", "Expert A", "Expert B"]


def test_expert_sheets_have_blank_score_cells(tmp_path):
    s = sampling.stratified_subsample(DF, frac=0.10)
    p = sampling.write_coding_workbook(s, tmp_path / "irr.xlsx")
    wb = openpyxl.load_workbook(p)
    ws = wb["Expert A"]
    header = [c.value for c in ws[1]]
    assert header[:4] == ["SR.NO", "YEAR", "TITLE", "ABSTRACT"]
    assert header[4:-1] == config.SCORE_COLS
    assert header[-1] == "Notes"
    for row in ws.iter_rows(min_row=2, min_col=5):
        assert all(c.value is None for c in row)


def test_codebook_sheet_lists_all_44_and_the_anchors(tmp_path):
    s = sampling.stratified_subsample(DF, frac=0.10)
    p = sampling.write_coding_workbook(s, tmp_path / "irr.xlsx")
    wb = openpyxl.load_workbook(p)
    values = [c.value for row in wb["Codebook"].iter_rows() for c in row if c.value]
    for code in config.SCORE_COLS:
        assert code in values
    assert any("Best-in-class" in str(v) for v in values)


def test_missing_title_is_written_as_a_blank_cell(tmp_path):
    """SR.NO 336 has no title in the source. openpyxl cannot serialise pd.NA,
    and the value must not become the string 'NA' either."""
    s = sampling.stratified_subsample(DF, frac=0.10)
    assert 336 in set(s["sr_no"]), "expected SR.NO 336 in the seeded sample"
    p = sampling.write_coding_workbook(s, tmp_path / "irr.xlsx")
    ws = openpyxl.load_workbook(p)["Expert A"]
    row = next(r for r in ws.iter_rows(min_row=2) if r[0].value == 336)
    assert row[2].value is None


def test_expert_sheets_have_validation_for_scores(tmp_path):
    """Score cells must have data validation restricting to 0-5 whole numbers."""
    s = sampling.stratified_subsample(DF, frac=0.10)
    p = sampling.write_coding_workbook(s, tmp_path / "irr.xlsx")
    wb = openpyxl.load_workbook(p)
    for expert in ("Expert A", "Expert B"):
        ws = wb[expert]
        assert len(ws.data_validations.dataValidation) >= 1, f"No validation found on {expert}"
        dv = ws.data_validations.dataValidation[0]
        # Check that the validation covers the score range (columns E onwards)
        assert "E2" in str(dv.sqref), f"Validation does not cover score columns on {expert}"
